from rest_framework import viewsets, permissions, status, filters
from django_filters.rest_framework import DjangoFilterBackend # 导入 DjangoFilterBackend
from rest_framework.response import Response
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

from .models import TestCase, TestCaseModule, Project, TestCaseScreenshot
from .serializers import TestCaseSerializer, TestCaseModuleSerializer, TestCaseScreenshotSerializer
from .permissions import IsProjectMemberForTestCase, IsProjectMemberForTestCaseModule
from .filters import TestCaseFilter # 导入自定义过滤器
# 确保导入项目自定义的权限类
from wharttest_django.permissions import HasModelPermission, permission_required

class TestCaseViewSet(viewsets.ModelViewSet):
    """
    用例视图集，处理用例的 CRUD 操作，并支持嵌套创建/更新用例步骤。
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/testcases/
    """
    serializer_class = TestCaseSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter] # 添加 DjangoFilterBackend
    filterset_class = TestCaseFilter # 使用自定义的 FilterSet
    search_fields = ['name', 'precondition']

    def get_permissions(self):
        """
        返回此视图所需权限的实例列表。
        这将覆盖 settings.DEFAULT_PERMISSION_CLASSES。
        """
        # 确保所有权限类都被实例化
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(), # 使用支持 @permission_required 装饰器的权限类
            IsProjectMemberForTestCase()
        ]

    def get_queryset(self):
        """
        根据 URL 中的 project_pk 过滤用例。
        确保只返回指定项目下的用例。
        """
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            # 权限类 IsProjectMemberForTestCase 已经检查了用户是否是此项目的成员
            # 所以这里可以直接返回项目下的用例
            return TestCase.objects.filter(project=project).select_related('creator', 'module').prefetch_related('steps')
        # 如果没有 project_pk (理论上不应该发生，因为路由是嵌套的)
        # 返回空 queryset 或根据需求抛出错误
        return TestCase.objects.none()

    def perform_create(self, serializer):
        """
        在创建用例时，自动关联项目和创建人。
        """
        project_pk = self.kwargs.get('project_pk')
        project = get_object_or_404(Project, pk=project_pk)
        # 权限类 IsProjectMemberForTestCase 已经确保用户是项目成员
        serializer.save(creator=self.request.user, project=project)

    # create 和 update 方法将使用序列化器中定义的嵌套写入逻辑。
    # DRF 的 ModelViewSet 会自动调用 serializer.save()，
    # 其中包含了处理嵌套 'steps' 的逻辑。

    # 如果需要更细致的控制，可以覆盖 create 和 update 方法，例如：
    # def create(self, request, *args, **kwargs):
    #     project_pk = self.kwargs.get('project_pk')
    #     project = get_object_or_404(Project, pk=project_pk)
    #
    #     # 可以在这里添加额外的逻辑，例如检查项目状态等
    #
    #     serializer = self.get_serializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)
    #     self.perform_create(serializer) # perform_create 中会设置 project 和 creator
    #     headers = self.get_success_headers(serializer.data)
    #     return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    # def update(self, request, *args, **kwargs):
    #     partial = kwargs.pop('partial', False)
    #     instance = self.get_object() # get_object 会进行对象级权限检查
    #
    #     # 可以在这里添加额外的逻辑
    #
    #     serializer = self.get_serializer(instance, data=request.data, partial=partial)
    #     serializer.is_valid(raise_exception=True)
    #     self.perform_update(serializer) # perform_update 默认只调用 serializer.save()
    #
    #     if getattr(instance, '_prefetched_objects_cache', None):
    #         # If 'prefetch_related' has been applied to a queryset, we need to
    #         # forcibly invalidate the prefetch cache on the instance.
    #         instance._prefetched_objects_cache = {}
    #
    #     return Response(serializer.data)

    # perform_update 默认调用 serializer.save()，我们的序列化器 update 方法会处理嵌套步骤。
    # perform_destroy 默认调用 instance.delete()。

    @action(detail=False, methods=['get', 'post'], url_path='export-excel')
    def export_excel(self, request, project_pk=None):
        """
        导出用例为Excel格式
        支持两种方式传递要导出的用例ID：
        1. GET请求通过ids参数: /api/projects/1/testcases/export-excel/?ids=1,2,3
        2. POST请求通过请求体: {"ids": [1, 2, 3]}
        如果不提供ids，则导出项目下所有用例
        """
        testcase_ids = None

        if request.method == 'POST':
            # POST请求，从请求体获取ids
            ids_data = request.data.get('ids', [])
            if ids_data:
                try:
                    testcase_ids = [int(id) for id in ids_data]
                except (ValueError, TypeError):
                    from rest_framework.response import Response
                    return Response(
                        {'error': 'ids参数格式错误，应为数字列表'},
                        status=400
                    )
        else:
            # GET请求，从查询参数获取ids
            ids_param = request.query_params.get('ids', '')
            if ids_param:
                try:
                    testcase_ids = [int(id.strip()) for id in ids_param.split(',') if id.strip()]
                except ValueError:
                    from rest_framework.response import Response
                    return Response(
                        {'error': 'ids参数格式错误，应为逗号分隔的数字列表'},
                        status=400
                    )

        # 根据是否提供了ids来过滤queryset
        if testcase_ids:
            queryset = self.get_queryset().filter(id__in=testcase_ids)
        else:
            queryset = self.get_queryset()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 设置表头
        headers = [
            '用例名称', '所属模块', '标签', '前置条件',
            '步骤描述', '预期结果', '编辑模式', '备注', '用例等级'
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row, testcase in enumerate(queryset, 2):
            # 获取模块路径
            module_path = self._get_module_path(testcase.module) if testcase.module else ""

            # 获取步骤描述和预期结果
            steps_desc, expected_results = self._format_steps(testcase.steps.all())

            # 写入数据行
            ws.cell(row=row, column=1, value=testcase.name)
            ws.cell(row=row, column=2, value=module_path)
            ws.cell(row=row, column=3, value="")  # 标签字段，当前数据库中没有
            ws.cell(row=row, column=4, value=testcase.precondition or "")
            ws.cell(row=row, column=5, value=steps_desc)
            ws.cell(row=row, column=6, value=expected_results)
            ws.cell(row=row, column=7, value="STEP")  # 编辑模式，固定为STEP
            ws.cell(row=row, column=8, value=testcase.notes or "")
            ws.cell(row=row, column=9, value=testcase.level)

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 创建HTTP响应
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # 获取项目名称用于文件名
        project = get_object_or_404(Project, pk=project_pk)
        filename = f"{project.name}_测试用例.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def _get_module_path(self, module):
        """
        获取模块的完整路径
        """
        if not module:
            return ""

        path_parts = []
        current = module
        while current:
            path_parts.insert(0, current.name)
            current = current.parent

        return "/" + "/".join(path_parts)

    def _format_steps(self, steps):
        """
        格式化步骤描述和预期结果
        """
        steps_desc = []
        expected_results = []

        for step in steps.order_by('step_number'):
            steps_desc.append(f"[{step.step_number}]{step.description}")
            expected_results.append(f"[{step.step_number}]{step.expected_result}")

        return "\n".join(steps_desc), "\n".join(expected_results)

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request, **kwargs):
        """
        批量删除用例
        POST请求体格式: {"ids": [1, 2, 3, 4]}
        """
        # 获取要删除的用例ID列表
        ids_data = request.data.get('ids', [])

        if not ids_data:
            return Response(
                {'error': '请提供要删除的用例ID列表'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 验证ID格式
        try:
            testcase_ids = [int(id) for id in ids_data]
        except (ValueError, TypeError):
            return Response(
                {'error': 'ids参数格式错误，应为数字列表'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not testcase_ids:
            return Response(
                {'error': '用例ID列表不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 获取当前项目下的用例queryset，确保数据隔离
        queryset = self.get_queryset()

        # 过滤出要删除的用例，确保只能删除当前项目下的用例
        testcases_to_delete = queryset.filter(id__in=testcase_ids)

        # 检查是否所有请求的ID都存在
        found_ids = list(testcases_to_delete.values_list('id', flat=True))
        not_found_ids = [id for id in testcase_ids if id not in found_ids]

        if not_found_ids:
            return Response(
                {
                    'error': f'以下用例ID不存在或不属于当前项目: {not_found_ids}',
                    'not_found_ids': not_found_ids
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 记录删除前的信息用于返回
        deleted_testcases_info = []
        for testcase in testcases_to_delete:
            deleted_testcases_info.append({
                'id': testcase.id,
                'name': testcase.name,
                'module': testcase.module.name if testcase.module else None
            })

        # 执行批量删除
        try:
            with transaction.atomic():
                # 删除用例（关联的步骤会因为外键级联删除而自动删除）
                deleted_count, deleted_details = testcases_to_delete.delete()

                return Response({
                    'message': f'成功删除 {len(deleted_testcases_info)} 个用例',
                    'deleted_count': len(deleted_testcases_info),
                    'deleted_testcases': deleted_testcases_info,
                    'deletion_details': deleted_details
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'删除过程中发生错误: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='upload-screenshots')
    @permission_required('testcases.add_testcasescreenshot')
    def upload_screenshots(self, request, project_pk=None, pk=None):
        """
        上传测试用例截屏（支持多张图片）
        POST /api/projects/{project_pk}/testcases/{pk}/upload-screenshots/
        请求体: multipart/form-data
        支持字段:
        - screenshots: 图片文件（可多个）
        - title: 图片标题（可选）
        - description: 图片描述（可选）
        - step_number: 对应步骤编号（可选）
        - mcp_session_id: MCP会话ID（可选）
        - page_url: 页面URL（可选）
        """
        testcase = self.get_object()

        # 获取上传的文件
        uploaded_files = request.FILES.getlist('screenshots')
        if not uploaded_files:
            # 兼容单文件上传
            if 'screenshot' in request.FILES:
                uploaded_files = [request.FILES['screenshot']]
            else:
                return Response(
                    {'error': '请提供截屏文件，字段名为 screenshots 或 screenshot'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # 验证文件数量限制
        if len(uploaded_files) > 10:
            return Response(
                {'error': '一次最多只能上传10张图片'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 验证文件类型和大小
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
        max_size = 5 * 1024 * 1024  # 5MB

        for file in uploaded_files:
            if file.content_type not in allowed_types:
                return Response(
                    {'error': f'文件 {file.name} 格式不支持，只支持 JPEG、PNG、GIF 格式'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if file.size > max_size:
                return Response(
                    {'error': f'文件 {file.name} 大小超过5MB限制'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            created_screenshots = []

            # 获取额外信息
            title = request.data.get('title', '')
            description = request.data.get('description', '')
            step_number = request.data.get('step_number')
            mcp_session_id = request.data.get('mcp_session_id', '')
            page_url = request.data.get('page_url', '')

            # 处理step_number
            if step_number:
                try:
                    step_number = int(step_number)
                except (ValueError, TypeError):
                    step_number = None

            # 为每个文件创建截屏记录
            for i, file in enumerate(uploaded_files):
                screenshot_data = {
                    'test_case': testcase.id,
                    'screenshot': file,
                    'title': f"{title} ({i+1})" if title and len(uploaded_files) > 1 else title,
                    'description': description,
                    'step_number': step_number,
                    'mcp_session_id': mcp_session_id,
                    'page_url': page_url,
                }

                serializer = TestCaseScreenshotSerializer(
                    data=screenshot_data,
                    context={'request': request}
                )

                if serializer.is_valid():
                    screenshot = serializer.save()
                    created_screenshots.append(serializer.data)
                else:
                    return Response(
                        {'error': f'文件 {file.name} 保存失败: {serializer.errors}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            return Response({
                'message': f'成功上传 {len(created_screenshots)} 张截屏',
                'screenshots': created_screenshots
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'上传失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='screenshots')
    def list_screenshots(self, request, project_pk=None, pk=None):
        """
        获取测试用例的所有截屏
        GET /api/projects/{project_pk}/testcases/{pk}/screenshots/
        """
        testcase = self.get_object()
        screenshots = testcase.screenshots.all()
        serializer = TestCaseScreenshotSerializer(
            screenshots,
            many=True,
            context={'request': request}
        )
        return Response(serializer.data)

    @action(detail=True, methods=['delete'], url_path='screenshots/(?P<screenshot_id>[^/.]+)')
    @permission_required('testcases.delete_testcasescreenshot')
    def delete_screenshot(self, request, project_pk=None, pk=None, screenshot_id=None):
        """
        删除指定的截屏
        DELETE /api/projects/{project_pk}/testcases/{pk}/screenshots/{screenshot_id}/
        """
        testcase = self.get_object()

        try:
            screenshot = testcase.screenshots.get(id=screenshot_id)
            screenshot.delete()
            return Response({
                'message': '截屏删除成功'
            }, status=status.HTTP_200_OK)
        except TestCaseScreenshot.DoesNotExist:
            return Response(
                {'error': '截屏不存在'},
                status=status.HTTP_404_NOT_FOUND
            )


class TestCaseModuleViewSet(viewsets.ModelViewSet):
    """
    用例模块视图集，处理模块的 CRUD 操作，支持5级子模块。
    API 端点将嵌套在项目下，例如 /api/projects/{project_pk}/testcase-modules/
    """
    serializer_class = TestCaseModuleSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']

    def get_permissions(self):
        """
        返回此视图所需权限的实例列表。
        """
        return [
            permissions.IsAuthenticated(),
            HasModelPermission(),
            IsProjectMemberForTestCaseModule()
        ]

    def get_queryset(self):
        """
        根据 URL 中的 project_pk 过滤模块。
        确保只返回指定项目下的模块。
        """
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            # 权限类 IsProjectMemberForTestCaseModule 已经检查了用户是否是此项目的成员
            return TestCaseModule.objects.filter(project=project).select_related('creator', 'parent')
        return TestCaseModule.objects.none()

    def perform_create(self, serializer):
        """
        在创建模块时，自动关联项目和创建人。
        """
        project_pk = self.kwargs.get('project_pk')
        project = get_object_or_404(Project, pk=project_pk)
        # 将项目实例添加到序列化器上下文，用于验证
        serializer.context['project'] = project
        # 保存模块，设置创建人和项目
        serializer.save(creator=self.request.user, project=project)

    def perform_destroy(self, instance):
        """
        删除模块前检查是否有关联的测试用例
        """
        if instance.testcases.exists():
            from rest_framework.exceptions import ValidationError
            testcase_count = instance.testcases.count()
            raise ValidationError(
                f"无法删除模块 '{instance.name}'，因为该模块下还有 {testcase_count} 个测试用例。请先删除或移动这些用例。"
            )
        instance.delete()

    def get_serializer_context(self):
        """
        为序列化器提供额外的上下文。
        """
        context = super().get_serializer_context()
        project_pk = self.kwargs.get('project_pk')
        if project_pk:
            project = get_object_or_404(Project, pk=project_pk)
            context['project'] = project
        return context
